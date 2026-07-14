from pathlib import Path

from openpyxl import load_workbook


VAT_FACTOR = 1.13
OUTPUT_DIR = Path("monthly_price_prediction_outputs") / "drop_limited_vars"
WORKBOOK = OUTPUT_DIR / "引入变量结果分析.xlsx"


def header_map(ws):
    return {cell.value: cell.column for cell in ws[1] if cell.value is not None}


def divide_numeric_columns(ws, columns):
    cols = header_map(ws)
    for name in columns:
        col = cols.get(name)
        if col is None:
            continue
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.value = cell.value / VAT_FACTOR


def replace_text(ws, replacements):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                text = cell.value
                for old, new in replacements.items():
                    text = text.replace(old, new)
                cell.value = text


def update_coefficients(wb):
    ws = wb["模型系数"]
    cols = header_map(ws)
    variable_col = cols["变量"]
    coef_col = cols["非标准化系数"]

    for row in range(2, ws.max_row + 1):
        variable = ws.cell(row=row, column=variable_col).value
        coef_cell = ws.cell(row=row, column=coef_col)
        if variable != "上月月均价" and isinstance(coef_cell.value, (int, float)):
            coef_cell.value = coef_cell.value / VAT_FACTOR


def rebuild_formulas(wb):
    coef_ws = wb["模型系数"]
    formula_ws = wb["模型公式"]
    coef_cols = header_map(coef_ws)
    formula_cols = header_map(formula_ws)

    by_material = {}
    for row in range(2, coef_ws.max_row + 1):
        material = coef_ws.cell(row=row, column=coef_cols["品种"]).value
        variable = coef_ws.cell(row=row, column=coef_cols["变量"]).value
        coef = coef_ws.cell(row=row, column=coef_cols["非标准化系数"]).value
        if material and variable and isinstance(coef, (int, float)):
            by_material.setdefault(material, []).append((variable, coef))

    for row in range(2, formula_ws.max_row + 1):
        material = formula_ws.cell(row=row, column=formula_cols["品种"]).value
        if material not in by_material:
            continue
        target = f"{material}_不含税月均价"
        parts = []
        intercept = 0.0
        for variable, coef in by_material[material]:
            if variable == "截距":
                intercept = coef
            else:
                display_var = "上月不含税月均价" if variable == "上月月均价" else variable
                sign = "+" if coef >= 0 else "-"
                parts.append(f" {sign} {abs(coef):.6f} * {display_var}")
        formula_ws.cell(row=row, column=formula_cols["因变量"]).value = target
        formula_ws.cell(row=row, column=formula_cols["公式"]).value = (
            f"{target} = {intercept:.6f}" + "".join(parts)
        )


def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)

    wb = load_workbook(WORKBOOK)
    train_ws = wb["训练测试集"]

    already_converted = False
    train_cols = header_map(train_ws)
    target_col = train_cols.get("目标变量")
    if target_col:
        first_target = train_ws.cell(row=2, column=target_col).value
        already_converted = isinstance(first_target, str) and "不含税" in first_target

    if already_converted:
        print(f"Already converted: {WORKBOOK}")
        return

    replace_text(
        train_ws,
        {
            "_月均价": "_不含税月均价",
            "上月月均价": "上月不含税月均价",
        },
    )

    divide_numeric_columns(wb["评估指标"], ["MAE", "RMSE", "Bias"])
    divide_numeric_columns(wb["测试集预测对比"], ["实际月均价", "预测月均价", "价格误差"])
    divide_numeric_columns(wb["未来月份预测"], ["预测月均价", "上月价格输入"])

    future_ws = wb["未来月份预测"]
    replace_text(
        future_ws,
        {
            "2026-06价格为实际月均价": "2026-06价格为实际不含税月均价",
            "上月月均价递推": "上月不含税月均价递推",
        },
    )

    update_coefficients(wb)
    rebuild_formulas(wb)

    wb.save(WORKBOOK)
    print(f"Converted to tax-exclusive prices: {WORKBOOK}")


if __name__ == "__main__":
    main()
